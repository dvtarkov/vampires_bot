# db/excel_templates.py
from __future__ import annotations
import asyncio
from dataclasses import dataclass
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from sqlalchemy import select
from db.session import get_session
from db.models import (
    User, District, Action, News, Politician,
    ControlLevel, ActionType, ActionStatus,
)


# --------------------------
# Вспомогательные структуры
# --------------------------

@dataclass
class Column:
    name: str
    required: bool
    typ: str
    example: Optional[str] = None
    note: Optional[str] = None


def _sheet_header(ws: Worksheet, title: str):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)


def _write_columns(ws: Worksheet, cols: list[Column]):
    """
    Формат шапки:
    R2: Описание/примечания (серая)
    R3: Названия полей (жёлтая)
    R4: Тип (светло-серая)
    Сами данные со строки 5.
    """
    fill_desc = PatternFill("solid", fgColor="ECECEC")
    fill_head = PatternFill("solid", fgColor="FFF2CC")
    fill_type = PatternFill("solid", fgColor="F3F3F3")

    # R2 notes / examples
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=i, value=c.note or (f"Пример: {c.example}" if c.example else ""))
        cell.alignment = Alignment(wrap_text=True)
        cell.fill = fill_desc

    # R3 names
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=3, column=i, value=(c.name + (" *" if c.required else "")))
        cell.font = Font(bold=True)
        cell.fill = fill_head

    # R4 types
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=4, column=i, value=c.typ)
        cell.fill = fill_type

    # Freeze header
    ws.freeze_panes = "A5"

    # Widths
    for i, c in enumerate(cols, start=1):
        width = max(12, len(c.name) + 2)
        if c.note:
            width = max(width, min(50, len(c.note) // 2))
        ws.column_dimensions[get_column_letter(i)].width = width


def _add_list_validation(ws: Worksheet, col_idx: int, items: Iterable[str], first_row: int = 5, last_row: int = 5000):
    lst = ",".join(items)
    dv = DataValidation(type="list", formula1=f'"{lst}"', allow_blank=True, showDropDown=True)
    addr = f"{get_column_letter(col_idx)}{first_row}:{get_column_letter(col_idx)}{last_row}"
    dv.add(addr)
    ws.add_data_validation(dv)


def _add_ref_validation(ws: Worksheet, col_idx: int, ref_sheet: str, ref_col_letter: str, first_row: int = 5,
                        last_row: int = 5000):
    # Валидация по диапазону на другом листе
    dv = DataValidation(
        type="list",
        formula1=f"='{ref_sheet}'!${ref_col_letter}$2:${ref_col_letter}$10000",
        allow_blank=True,
        showDropDown=True,
    )
    addr = f"{get_column_letter(col_idx)}{first_row}:{get_column_letter(col_idx)}{last_row}"
    dv.add(addr)
    ws.add_data_validation(dv)


# --------------------------
# Основной экспортёр
# --------------------------

async def export_excel_templates(path: str) -> str:
    """
    Генерирует XLSX со вкладками для Users, Districts, Actions, News, Politicians,
    справочниками для валидации и ПРЕДЗАПОЛНЕНИЕМ текущих данных БД
    (Users, Districts, Politicians). Возвращает путь к файлу.
    """
    from datetime import datetime

    def _iso(dt: Optional[datetime]) -> Optional[str]:
        if not dt:
            return None
        s = dt.isoformat()
        # чуть красивее ISO для UTC
        return s.replace("+00:00", "Z")

    def _enum_name(v):
        return getattr(v, "name", v)

    wb = Workbook()
    wb.remove(wb.active)

    # ---- Справочники (для валидаций) ----
    async with get_session() as session:
        districts = (await session.execute(
            select(District.id, District.name).order_by(District.name)
        )).all()
        actions = (await session.execute(
            select(Action.id, Action.title, Action.kind).order_by(Action.id.desc())
        )).all()

    # REF_Districts
    ws_ref_d = wb.create_sheet("REF_Districts")
    _sheet_header(ws_ref_d, "Справочник районов (не редактируйте имена столбцов)")
    ws_ref_d["A2"] = "id"; ws_ref_d["B2"] = "name"
    ws_ref_d["A2"].font = ws_ref_d["B2"].font = Font(bold=True)
    for i, (did, dname) in enumerate(districts, start=3):
        ws_ref_d.cell(row=i, column=1, value=did)
        ws_ref_d.cell(row=i, column=2, value=dname)
    ws_ref_d.freeze_panes = "A3"
    ws_ref_d.column_dimensions["A"].width = 10
    ws_ref_d.column_dimensions["B"].width = 40

    # REF_Actions
    ws_ref_a = wb.create_sheet("REF_Actions")
    _sheet_header(ws_ref_a, "Справочник действий (не редактируйте имена столбцов)")
    ws_ref_a["A2"] = "id"; ws_ref_a["B2"] = "title"; ws_ref_a["C2"] = "kind"
    for c in ("A2","B2","C2"): ws_ref_a[c].font = Font(bold=True)
    for i, (aid, title, kind) in enumerate(actions, start=3):
        ws_ref_a.cell(row=i, column=1, value=aid)
        ws_ref_a.cell(row=i, column=2, value=title or "")
        ws_ref_a.cell(row=i, column=3, value=kind)
    ws_ref_a.freeze_panes = "A3"
    for col, w in zip(("A","B","C"), (10, 40, 18)):
        ws_ref_a.column_dimensions[col].width = w

    # ---- Users ----
    ws_users = wb.create_sheet("Users")
    _sheet_header(ws_users, "Users — импорт поддерживает upsert по tg_id")
    users_cols = [
        Column("tg_id", True, "int", "305632047", "Уникальный Telegram ID (обязательно)"),
        Column("username", False, "str", "john_doe"),
        Column("first_name", False, "str", "John"),
        Column("last_name", False, "str", "Doe"),
        Column("in_game_name", False, "str", "Vamp#123"),
        Column("language_code", False, "str", "ru"),
        Column("money", False, "int", "0"),
        Column("influence", False, "int", "0"),
        Column("information", False, "int", "0"),
        Column("force", False, "int", "0"),
        Column("ideology", False, "int [-5..+5]", "0"),
        Column("faction", False, "str", "Regime"),
        Column("available_actions", False, "int", "0"),
        Column("max_available_actions", False, "int", "5"),
        Column("actions_refresh_at", False, "datetime ISO8601", "2025-08-14T19:00:00Z"),
    ]
    _write_columns(ws_users, users_cols)

    # ---- Districts ----
    ws_d = wb.create_sheet("Districts")
    _sheet_header(ws_d, "Districts — импорт: create/update по id (если указан)")
    d_cols = [
        Column("id", False, "int", "1", "Оставьте пустым для создания"),
        Column("name", True, "str", "Stari Grad"),
        Column("owner_tg_id", True, "int", "305632047", "Владелец — tg_id пользователя"),
        Column("control_points", False, "int", "0"),
        Column("control_level", False, "enum", "MINIMAL", "Выпадающий список"),
        Column("resource_multiplier", False, "float", "0.4"),
        Column("base_money", False, "int", "100"),
        Column("base_influence", False, "int", "10"),
        Column("base_information", False, "int", "5"),
        Column("base_force", False, "int", "0"),
    ]
    _write_columns(ws_d, d_cols)
    _add_list_validation(ws_d, col_idx=5, items=[e.name for e in ControlLevel])

    # ---- Actions (шаблон + валидации, без предзаполнения) ----
    ws_a = wb.create_sheet("Actions")
    _sheet_header(ws_a, "Actions — создание/обновление; FK через выпадающие списки")
    a_cols = [
        Column("id", False, "int", "1", "Оставьте пустым для создания"),
        Column("owner_tg_id", True, "int", "305632047", "Владелец — tg_id пользователя"),
        Column("kind", True, "str", "defend"),
        Column("title", False, "str", "Оборона Старого Града"),
        Column("status", False, "enum", "pending", "Выпадающий список"),
        Column("district_id", False, "ref District", "1", "Список с REF_Districts"),
        Column("type", False, "enum", "individual", "Выпадающий список"),
        Column("parent_action_id", False, "ref Action", "", "Для экшенов-помощи/детей"),
        Column("force", False, "int", "0"),
        Column("money", False, "int", "0"),
        Column("influence", False, "int", "0"),
        Column("information", False, "int", "0"),
    ]
    _write_columns(ws_a, a_cols)
    _add_list_validation(ws_a, 5, [e.value for e in ActionStatus])  # status
    _add_ref_validation(ws_a, 6, "REF_Districts", "A")               # district_id by ids
    _add_list_validation(ws_a, 7, [e.value for e in ActionType])     # type
    _add_ref_validation(ws_a, 8, "REF_Actions", "A")                 # parent_action_id

    # ---- News (шаблон) ----
    ws_n = wb.create_sheet("News")
    _sheet_header(ws_n, "News — media_urls через JSON-массив строк")
    n_cols = [
        Column("id", False, "int", "", "Оставьте пустым для создания"),
        Column("title", True, "str", "Заголовок"),
        Column("body", True, "text", "Большой текст новости"),
        Column("media_urls", False, "json list[str]", '["https://site/img1.jpg"]', "JSON-массив ссылок"),
        Column("action_id", False, "ref Action", "", "Выпадающий список по id"),
    ]
    _write_columns(ws_n, n_cols)
    _add_ref_validation(ws_n, 5, "REF_Actions", "A")

    # ---- Politicians ----
    ws_p = wb.create_sheet("Politicians")
    _sheet_header(ws_p, "Politicians")
    p_cols = [
        Column("id", False, "int", "", "Оставьте пустым для создания"),
        Column("name", True, "str", "Слободан Милошевич"),
        Column("role_and_influence", True, "text", "Глава государства, контроль над госаппаратом"),
        Column("district_id", False, "ref District", "", "Выпадающий список по id"),
        Column("ideology", False, "int [-5..+5]", "0"),
        Column("influence", False, "int (±)", "0"),
        Column("bonuses_penalties", False, "text", "+5 ОК за каждую заявку ..."),
    ]
    _write_columns(ws_p, p_cols)
    _add_ref_validation(ws_p, 4, "REF_Districts", "A")

    # ===== Предзаполнение текущими данными БД =====
    # ===== Предзаполнение текущими данными БД (без загрузки ORM-объектов) =====
    async with get_session() as session:
        # Users → выбираем только нужные поля, чтобы не триггерить загрузку связей
        users_rows = (await session.execute(
            select(
                User.tg_id, User.username, User.first_name, User.last_name,
                User.in_game_name, User.language_code,
                User.money, User.influence, User.information, User.force,
                User.ideology, User.faction,
                User.available_actions, User.max_available_actions,
                User.actions_refresh_at,
            ).order_by(User.id)
        )).all()

        for r, row in enumerate(users_rows, start=5):
            (tg_id, username, first_name, last_name, in_game_name, language_code,
             money, influence, information, force, ideology, faction,
             available_actions, max_available_actions, actions_refresh_at) = row
            ws_users.cell(r, 1, tg_id)
            ws_users.cell(r, 2, username)
            ws_users.cell(r, 3, first_name)
            ws_users.cell(r, 4, last_name)
            ws_users.cell(r, 5, in_game_name)
            ws_users.cell(r, 6, language_code)
            ws_users.cell(r, 7, money)
            ws_users.cell(r, 8, influence)
            ws_users.cell(r, 9, information)
            ws_users.cell(r, 10, force)
            ws_users.cell(r, 11, ideology)
            ws_users.cell(r, 12, faction)
            ws_users.cell(r, 13, available_actions)
            ws_users.cell(r, 14, max_available_actions)
            ws_users.cell(r, 15, actions_refresh_at.isoformat().replace("+00:00", "Z") if actions_refresh_at else None)

        # Districts → берём owner_tg_id через LEFT JOIN, только существующие колонки
        districts_rows = (await session.execute(
            select(
                District.id, District.name,
                User.tg_id.label("owner_tg_id"),
                District.control_points, District.control_level,
                District.resource_multiplier,
                District.base_money, District.base_influence,
                District.base_information, District.base_force,
            )
            .select_from(District)
            .join(User, District.owner_id == User.id, isouter=True)
            .order_by(District.id)
        )).all()

        for r, row in enumerate(districts_rows, start=5):
            (did, dname, owner_tg_id, control_points, control_level,
             resource_multiplier, base_money, base_influence,
             base_information, base_force) = row
            ws_d.cell(r, 1, did)
            ws_d.cell(r, 2, dname)
            ws_d.cell(r, 3, owner_tg_id)
            ws_d.cell(r, 4, control_points)
            ws_d.cell(r, 5, getattr(control_level, "name", control_level))  # enum → NAME
            ws_d.cell(r, 6, resource_multiplier)
            ws_d.cell(r, 7, base_money)
            ws_d.cell(r, 8, base_influence)
            ws_d.cell(r, 9, base_information)
            ws_d.cell(r, 10, base_force)

        # Politicians → только нужные поля
        pol_rows = (await session.execute(
            select(
                Politician.id, Politician.name, Politician.role_and_influence,
                Politician.district_id, Politician.ideology,
                Politician.influence, Politician.bonuses_penalties,
            ).order_by(Politician.id)
        )).all()

        for r, row in enumerate(pol_rows, start=5):
            (pid, name, role_and_infl, district_id, ideology, influence, bp) = row
            ws_p.cell(r, 1, pid)
            ws_p.cell(r, 2, name)
            ws_p.cell(r, 3, role_and_infl)
            ws_p.cell(r, 4, district_id)
            ws_p.cell(r, 5, ideology)
            ws_p.cell(r, 6, influence)
            ws_p.cell(r, 7, bp)

    # Сохраняем
    wb.save(path)
    return path



# Для локального запуска:
if __name__ == "__main__":
    import os

    out = os.path.abspath("game_models_template.xlsx")
    print("Writing:", out)
    asyncio.run(export_excel_templates(out))
    print("Done.")
