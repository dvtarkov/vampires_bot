# db/excel_import.py
from __future__ import annotations

import json
import math
from typing import Any, Dict, Iterable, Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.exc import SQLAlchemyError

from db.session import get_session
from db.models import (
    User, District, Action, News, Politician,
    ControlLevel, ActionType, ActionStatus,
)


# ---------- утилиты приведения типов ----------

def _is_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    # openpyxl иногда даёт float("nan")
    if isinstance(v, float) and math.isnan(v):
        return True
    if isinstance(v, str) and v.strip().lower() in {"nan", "none", "null"}:
        return True
    return False

def _to_str_or_none(v: Any) -> Optional[str]:
    return None if _is_empty(v) else str(v)

def _to_int_or_none(v: Any) -> Optional[int]:
    if _is_empty(v):
        return None
    try:
        if isinstance(v, bool):
            return int(v)
        return int(float(v))  # на случай "1.0"
    except (ValueError, TypeError):
        return None

def _to_int_or_zero(v: Any) -> int:
    iv = _to_int_or_none(v)
    return 0 if iv is None else iv

def _to_float_or_none(v: Any) -> Optional[float]:
    if _is_empty(v):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def _to_float_or_zero(v: Any) -> float:
    fv = _to_float_or_none(v)
    return 0.0 if fv is None else fv

def _enum_by_value(enum_cls, v: Any, default):
    if _is_empty(v):
        return default
    s = str(v).strip()
    # поддержим как .name, так и .value
    for item in enum_cls:
        if s == item.name or s == item.value:
            return item
    return default

def _json_list_of_str_or_empty(v: Any) -> list[str]:
    if _is_empty(v):
        return []
    if isinstance(v, list):
        return [str(x) for x in v]
    if isinstance(v, str):
        try:
            data = json.loads(v)
            if isinstance(data, list):
                return [str(x) for x in data]
        except Exception:
            # если дали одну ссылку строкой — тоже примем
            return [v]
    return []


# ---------- чтение таблиц (данные с R5) ----------

def _iter_rows(ws, want_cols: Iterable[str]) -> Iterable[Dict[str, Any]]:
    """
    ws: лист, где R3 — имена колонок, R4 — типы, данные c R5.
    want_cols: имена колонок, которые ожидаем (в любом порядке).
    """
    # карта "имя поля" -> индекс колонки
    headers: Dict[str, int] = {}
    col_idx = 1
    while True:
        name_cell = ws.cell(row=3, column=col_idx)
        if _is_empty(name_cell.value) and col_idx > 1000:
            break
        name = str(name_cell.value or "").strip().replace(" *", "")
        if not name:
            # дошли до пустой колонки — останавливаемся
            break
        headers[name] = col_idx
        col_idx += 1

    # проверка что нужные колонки вообще есть (мягко)
    missing = [c for c in want_cols if c not in headers]
    if missing:
        # не падаем: просто будем возвращать None для отсутствующих
        pass

    row = 5
    while True:
        # критерий окончания — полностью пустая строка (по всем известным колонкам)
        if all(_is_empty(ws.cell(row=row, column=headers.get(c, 10**9)).value) for c in headers.keys()):
            # чтобы не обрубить данные, проверим ещё 3 пустых подряд
            empty_block = True
            for k in range(1, 4):
                if any(not _is_empty(ws.cell(row=row+k, column=i).value) for i in headers.values()):
                    empty_block = False
                    break
            if empty_block:
                break

        record = {}
        for key, idx in headers.items():
            record[key] = ws.cell(row=row, column=idx).value
        yield record
        row += 1


# ---------- импорт конкретных сущностей ----------

async def _import_users(ws) -> int:
    """
    Upsert по tg_id.
    Возвращает число обработанных строк.
    """
    processed = 0
    async with get_session() as session:
        async for _ in _aiter(_iter_rows(ws, (
            "tg_id","username","first_name","last_name","in_game_name","language_code",
            "money","influence","information","force",
            "ideology","faction","available_actions","max_available_actions","actions_refresh_at",
        ))):
            pass  # для type-checker =)

        for row in _iter_rows(ws, (
            "tg_id","username","first_name","last_name","in_game_name","language_code",
            "money","influence","information","force",
            "ideology","faction","available_actions","max_available_actions","actions_refresh_at",
        )):
            tg_id = _to_int_or_none(row.get("tg_id"))
            if tg_id is None:
                continue  # без tg_id не импортируем
            # подготовка значений
            values = {
                "username": _to_str_or_none(row.get("username")),
                "first_name": _to_str_or_none(row.get("first_name")),
                "last_name": _to_str_or_none(row.get("last_name")),
                "in_game_name": _to_str_or_none(row.get("in_game_name")),
                "language_code": _to_str_or_none(row.get("language_code")),
                "money": _to_int_or_zero(row.get("money")),
                "influence": _to_int_or_zero(row.get("influence")),
                "information": _to_int_or_zero(row.get("information")),
                "force": _to_int_or_zero(row.get("force")),
                "ideology": _to_int_or_zero(row.get("ideology")),
                "faction": _to_str_or_none(row.get("faction")),
                "available_actions": _to_int_or_zero(row.get("available_actions")),
                "max_available_actions": _to_int_or_zero(row.get("max_available_actions")),
                # actions_refresh_at — оставляем как есть (строкой не пишем),
                # т.к. в шаблоне это ISO; если нужно — допиши парсер ISO → datetime.
            }

            existing = (await session.execute(select(User).where(User.tg_id == tg_id))).scalars().first()
            if existing:
                await User.update_by_tg_id(session, tg_id, **values)
            else:
                await User.create(session, tg_id=tg_id, **values)
            processed += 1
    return processed


async def _import_districts(ws) -> int:
    processed = 0
    async with get_session() as session:
        for row in _iter_rows(ws, (
            "id","name","owner_tg_id","control_points","control_level",
            "resource_multiplier","base_money","base_influence","base_information","base_force",
        )):
            print(row)
            name = _to_str_or_none(row.get("name"))
            if not name:
                continue

            owner_tg_id = _to_int_or_none(row.get("owner_tg_id"))
            if owner_tg_id is None:
                continue
            owner = (await session.execute(select(User).where(User.tg_id == owner_tg_id))).scalars().first()
            if not owner:
                # создадим «каркасного» юзера
                owner = await User.create(session, tg_id=owner_tg_id)

            did = _to_int_or_none(row.get("id"))
            values = {
                "name": name,
                "owner_id": owner.id,
                "control_points": _to_int_or_zero(row.get("control_points")),
                "control_level": _enum_by_value(ControlLevel, row.get("control_level"), ControlLevel.MINIMAL),
                "resource_multiplier": _to_float_or_zero(row.get("resource_multiplier") or 0.4),
                "base_money": _to_int_or_zero(row.get("base_money")),
                "base_influence": _to_int_or_zero(row.get("base_influence")),
                "base_information": _to_int_or_zero(row.get("base_information")),
                "base_force": _to_int_or_zero(row.get("base_force")),
            }

            if did:
                obj = await District.get_by_id(session, did)
                if obj:
                    # прямое обновление полей
                    obj.name = values["name"]
                    obj.owner_id = values["owner_id"]
                    obj.control_points = values["control_points"]
                    obj.control_level = values["control_level"]
                    obj.resource_multiplier = values["resource_multiplier"]
                    obj.base_money = values["base_money"]
                    obj.base_influence = values["base_influence"]
                    obj.base_information = values["base_information"]
                    obj.base_force = values["base_force"]
                    await session.commit()
                else:
                    await District.create(session, **values)
            else:
                await District.create(session, **values)
            processed += 1
    return processed


async def _import_actions(ws) -> int:
    processed = 0
    async with get_session() as session:
        for row in _iter_rows(ws, (
            "id","owner_tg_id","kind","title","status","district_id","type","parent_action_id",
            "force","money","influence","information",
        )):
            owner_tg_id = _to_int_or_none(row.get("owner_tg_id"))
            kind = _to_str_or_none(row.get("kind"))
            if owner_tg_id is None or not kind:
                continue

            owner = (await session.execute(select(User).where(User.tg_id == owner_tg_id))).scalars().first()
            if not owner:
                owner = await User.create(session, tg_id=owner_tg_id)

            aid = _to_int_or_none(row.get("id"))
            values = {
                "owner_id": owner.id,
                "kind": kind,
                "title": _to_str_or_none(row.get("title")),
                "status": _enum_by_value(ActionStatus, row.get("status"), ActionStatus.PENDING),
                "district_id": _to_int_or_none(row.get("district_id")),
                "type": _enum_by_value(ActionType, row.get("type"), ActionType.INDIVIDUAL),
                "parent_action_id": _to_int_or_none(row.get("parent_action_id")),
                "force": _to_int_or_zero(row.get("force")),
                "money": _to_int_or_zero(row.get("money")),
                "influence": _to_int_or_zero(row.get("influence")),
                "information": _to_int_or_zero(row.get("information")),
            }

            if aid:
                obj = await Action.get_by_id(session, aid)
                if obj:
                    obj.owner_id = values["owner_id"]
                    obj.kind = values["kind"]
                    obj.title = values["title"]
                    obj.status = values["status"]
                    obj.district_id = values["district_id"]
                    obj.type = values["type"]
                    obj.parent_action_id = values["parent_action_id"]
                    obj.force = values["force"]
                    obj.money = values["money"]
                    obj.influence = values["influence"]
                    obj.information = values["information"]
                    await session.commit()
                else:
                    await Action.create(session, **values)
            else:
                await Action.create(session, **values)
            processed += 1
    return processed


async def _import_news(ws) -> int:
    processed = 0
    async with get_session() as session:
        for row in _iter_rows(ws, ("id","title","body","media_urls","action_id")):
            title = _to_str_or_none(row.get("title"))
            body = _to_str_or_none(row.get("body"))
            if not title or not body:
                continue

            nid = _to_int_or_none(row.get("id"))
            values = {
                "title": title,
                "body": body,
                "media_urls": _json_list_of_str_or_empty(row.get("media_urls")),
                "action_id": _to_int_or_none(row.get("action_id")),
            }

            if nid:
                await News.update(session, nid, **values)
            else:
                await News.create(session, **values)
            processed += 1
    return processed


async def _import_politicians(ws) -> int:
    processed = 0
    async with get_session() as session:
        for row in _iter_rows(ws, (
            "id","name","role_and_influence","ideology","influence","bonuses_penalties",
        )):
            name = _to_str_or_none(row.get("name"))
            role = _to_str_or_none(row.get("role_and_influence"))
            if not name or not role:
                continue

            pid = _to_int_or_none(row.get("id"))
            values = {
                "name": name,
                "role_and_influence": role,
                "district_id": _to_int_or_none(row.get("district_id")),
                "ideology": max(-5, min(5, _to_int_or_zero(row.get("ideology")))),
                "influence": _to_int_or_zero(row.get("influence")),
                "bonuses_penalties": _to_str_or_none(row.get("bonuses_penalties")),
            }

            if pid:
                await Politician.update(session, pid, **values)
            else:
                await Politician.create(session, **values)
            processed += 1
    return processed


# ---------- основной вход ----------

async def import_excel(path: str) -> dict[str, int]:
    """
    Импортирует все листы из XLSX в базу.
    Возвращает счётчики по листам.
    """
    wb = load_workbook(filename=path, data_only=True)

    counters = {
        "Users": 0,
        "Districts": 0,
        "Actions": 0,
        "News": 0,
        "Politicians": 0,
    }

    try:
        if "Users" in wb.sheetnames:
            counters["Users"] = await _import_users(wb["Users"])
        if "Districts" in wb.sheetnames:
            counters["Districts"] = await _import_districts(wb["Districts"])
        if "Actions" in wb.sheetnames:
            counters["Actions"] = await _import_actions(wb["Actions"])
        if "News" in wb.sheetnames:
            counters["News"] = await _import_news(wb["News"])
        if "Politicians" in wb.sheetnames:
            counters["Politicians"] = await _import_politicians(wb["Politicians"])
    except SQLAlchemyError as e:
        # при желании можно логгировать и откатывать на уровне отдельных сессий
        raise
    finally:
        wb.close()

    return counters


# Небольшой helper для совместимости с синхронным генератором
async def _aiter(it):
    for x in it:
        yield x


if __name__ == "__main__":
    import asyncio, os
    p = os.path.abspath("game_models_template.xlsx")
    print("Importing:", p)
    print(asyncio.run(import_excel(p)))
