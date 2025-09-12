# game_cycle.py
"""
Запуск:
    export DATABASE_URL="sqlite+aiosqlite:///./game.db"  # или ваша строка подключения
    export COMBAT_RATES_PATH="./config/combat_rates.json"
    python game_cycle.py

Особенности:
- SUPPORT-экшены агрегируются в parent.
- Ресурсы экшенов конвертируются в очки атаки/обороны по JSON-курсу.
- "on point" даёт +20 очков, но при одновременных on_point атаке и обороне район становится «спорным».
- Множитель ресурсов района зависит от близости идеологий владельца и «прикреплённого» политика
  и квантуется шагом 0.1 в диапазоне [0.4..1.2].
- Оставшаяся оборона после атак → в control_points района.
- Базовая оборона на цикл формируется из control_points района.
- Новости пишутся в XLSX (по UTC-таймстампу цикла). Уведомления игрокам — через бота (если доступен).
- Добавлено подробное логирование всех шагов.
"""

import asyncio
import json
import logging
import os
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select, update
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.orm import sessionmaker

from db.models import (
    Base,
    User,
    District,
    Action,
    Politician,
    ActionStatus,
    ActionType,
    user_scouts_districts,
)
from services.notify import notify_user
from utils.raw_body_input import add_raw_row


# ===========================
#    Константы/настройки
# ===========================
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite+aiosqlite:///./game.db")
COMBAT_RATES_PATH = os.getenv("COMBAT_RATES_PATH", "./config/combat_rates.json")

ATTACK_KIND = "attack"
DEFENSE_KIND = "defend"
SCOUT_KINDS = {"scout_district", "scout_info"}

ORDER_ATTACKS_ASC = True  # порядок атак по created_at

# XLSX-выгрузка новостей
CYCLE_TS: Optional[str] = None
CYCLE_XLSX_PATH: Optional[Path] = None
NEWS_HEADERS = ["created_at_utc", "tag", "title", "body", "action_id", "district_id"]

# ===========================
#          Логгер
# ===========================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
log = logging.getLogger("game_cycle")


class StepTimer:
    """Простой контекстный менеджер для логирования длительности шага."""

    def __init__(self, name: str, level: int = logging.INFO):
        self.name = name
        self.level = level
        self._start = 0.0

    def __enter__(self):
        self._start = time.perf_counter()
        log.log(self.level, f"▶ {self.name} — старт")
        return self

    def __exit__(self, exc_type, exc, tb):
        dur = time.perf_counter() - self._start
        if exc_type is None:
            log.log(self.level, f"✓ {self.name} — завершено за {dur:.3f}s")
        else:
            log.exception(f"✗ {self.name} — ERROR через {dur:.3f}s")
        # не подавляем исключение
        return False


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _quantize_tenth(x: float, lo: float = 0.40, hi: float = 1.20) -> float:
    x = max(lo, min(hi, x))
    return float(Decimal(str(x)).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP))


def _ensure_sheet(wb: Workbook, name: str, headers: List[str]) -> None:
    if name not in wb.sheetnames:
        ws = wb.create_sheet(title=name)
        ws.append(headers)


def _ensure_cycle_workbook() -> Path:
    """Создаёт (при необходимости) XLSX для текущего цикла и возвращает путь к нему."""
    global CYCLE_XLSX_PATH
    if not CYCLE_TS:
        raise RuntimeError("CYCLE_TS не задан. Устанавливается в начале run_game_cycle().")
    path = Path(f"./exports/{CYCLE_TS}.xlsx")
    if not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "news"
        ws.append(NEWS_HEADERS)
        wb.save(path)
        log.info(f"Создан файл отчёта цикла: {path}")
    CYCLE_XLSX_PATH = path
    return path


@dataclass(frozen=True)
class CombatRates:
    attack: Dict[str, float]
    defense: Dict[str, float]
    on_point_bonus: int = 20

    @staticmethod
    def load(path: str) -> "CombatRates":
        p = Path(path)
        if not p.exists():
            raise FileNotFoundError(f"Combat rates file not found: {p}")
        data = json.loads(p.read_text(encoding="utf-8"))
        log.info("Загружены курсы конверсии из %s", p)
        log.debug("Курсы атаки: %s", data.get("attack"))
        log.debug("Курсы обороны: %s", data.get("defense"))
        return CombatRates(
            attack=data.get("attack", {}),
            defense=data.get("defense", {}),
            on_point_bonus=int(data.get("on_point_bonus", 20)),
        )


# ===========================
#     NEWS → XLSX helpers
# ===========================
async def add_news(
    session: AsyncSession,
    title: str,
    body: str,
    action_id: Optional[int] = None,
    *,
    district_id: Optional[int] = None,
    tag: str = "auto generated",
):
    """Пишет строку новости в общий лист 'news' и (если district_id задан) в лист района."""
    try:
        path = _ensure_cycle_workbook()
        wb = load_workbook(path)

        # общий лист
        _ensure_sheet(wb, "news", NEWS_HEADERS)
        row = [now_utc().isoformat(), tag, title, body, action_id, district_id]
        wb["news"].append(row)

        # лист конкретного района
        if district_id is not None:
            sheet_name = f"district_{district_id}"
            _ensure_sheet(wb, sheet_name, NEWS_HEADERS)
            wb[sheet_name].append(row)

        wb.save(path)
        log.debug("NEWS@%s: %s | %s", district_id or "-", title, body[:120].replace("\n", " "))
    except Exception:
        log.exception("Не удалось записать новость в XLSX")


# ===========================
#   SUPPORT → PARENT AGGREG
# ===========================
async def aggregate_supports(session: AsyncSession) -> Tuple[List[int], List[int]]:
    """
    Суммирует все SUPPORT-действия (status=PENDING) в их parent_action.
    Возвращает:
        (список id обработанных SUPPORT, список id parent, в которые агрегировали)
    """
    with StepTimer("Агрегация SUPPORT-экшенов"):
        stmt = select(Action).where(
            Action.status == ActionStatus.PENDING,
            Action.parent_action_id.is_not(None),
            Action.type == ActionType.SUPPORT,
        )
        res = await session.execute(stmt)
        supports: List[Action] = list(res.scalars().all())
        log.info("Найдено SUPPORT-экшенов: %d", len(supports))

        if not supports:
            return [], []

        # Группируем по parent_action_id
        by_parent: Dict[int, List[Action]] = defaultdict(list)
        for s in supports:
            if s.parent_action_id:
                by_parent[s.parent_action_id].append(s)

        parent_ids: List[int] = list(by_parent.keys())

        q = await session.execute(select(Action).where(Action.id.in_(parent_ids)))
        parents: Dict[int, Action] = {a.id: a for a in q.scalars().all()}
        log.debug("Родителей для агрегации: %d", len(parents))

        processed_support_ids: List[int] = []
        touched_parents: List[int] = []

        for pid, group in by_parent.items():
            parent = parents.get(pid)
            if not parent or parent.status != ActionStatus.PENDING:
                for s in group:
                    processed_support_ids.append(s.id)
                log.debug("Пропущен parent %s (нет/не PENDING), помечаем supports как DONE: %s", pid, [s.id for s in group])
                continue

            total_force = sum(max(0, int(s.force)) for s in group)
            total_money = sum(max(0, int(s.money)) for s in group)
            total_infl = sum(max(0, int(s.influence)) for s in group)
            total_info = sum(max(0, int(s.information)) for s in group)

            parent.force += total_force
            parent.money += total_money
            parent.influence += total_infl
            parent.information += total_info

            log.debug(
                "Parent %s пополнен: +F=%s +M=%s +I=%s +Info=%s",
                pid, total_force, total_money, total_infl, total_info
            )

            touched_parents.append(parent.id)
            processed_support_ids.extend([s.id for s in group])

        await session.commit()

        if processed_support_ids:
            await session.execute(
                update(Action)
                .where(Action.id.in_(processed_support_ids))
                .values(status=ActionStatus.DONE, updated_at=now_utc())
            )
            await session.commit()

        log.info("SUPPORT обработано: %d; родителей затронуто: %d", len(processed_support_ids), len(touched_parents))
        return processed_support_ids, touched_parents


# ===========================
#   RESOURCES → POINTS
# ===========================
def resources_to_points(kind: str, action: Action, rates: CombatRates) -> int:
    """Конвертирует ресурсы экшена в очки (учёт on_point)."""
    table = rates.attack if kind == ATTACK_KIND else rates.defense
    force_w = float(table.get("force", 0))
    money_w = float(table.get("money", 0))
    infl_w = float(table.get("influence", 0))
    info_w = float(table.get("information", 0))

    pts = (
        max(0, int(action.force)) * force_w
        + max(0, int(action.money)) * money_w
        + max(0, int(action.influence)) * infl_w
        + max(0, int(action.information)) * info_w
    )
    pts = int(round(pts))
    if action.on_point:
        pts += rates.on_point_bonus
    return max(0, pts)


# ===========================
#  CONTESTED DETECTION
# ===========================
async def detect_contested_districts(session: AsyncSession) -> List[int]:
    """
    Спорные районы:
      • есть >=2 атак с moving_on_point, ИЛИ
      • есть >=1 атака и >=1 защита с moving_on_point.
    Дополнительно: рассылает авторам on-point действий вопрос "Вы победили?".
    """
    with StepTimer("Поиск спорных районов (moving_on_point + уведомления)"):
        # moving_on_point — если поля нет, используем on_point как фолбэк
        ONP = getattr(Action, "moving_on_point", Action.on_point)

        # Забираем все pending on-point атаки/защиты по районам
        stmt = (
            select(Action.id, Action.owner_id, Action.district_id, Action.kind)
            .where(
                Action.status == ActionStatus.PENDING,
                Action.district_id.is_not(None),
                ONP.is_(True),
                Action.kind.in_([ATTACK_KIND, DEFENSE_KIND]),
            )
        )
        res = await session.execute(stmt)
        rows: list[tuple[int, int, int, str]] = list(res.all())

        # Разложим по районам
        by_district: dict[int, dict[str, list[tuple[int, int]]]] = defaultdict(lambda: {"attack": [], "defense": []})
        for a_id, owner_id, district_id, kind in rows:
            did = int(district_id)
            if str(kind) == ATTACK_KIND:
                by_district[did]["attack"].append((int(a_id), int(owner_id)))
            else:
                by_district[did]["defense"].append((int(a_id), int(owner_id)))

        # Правила спорности
        contested = []
        for did, kinds in by_district.items():
            a_cnt = len(kinds["attack"])
            d_cnt = len(kinds["defense"])
            if a_cnt >= 2 or (a_cnt >= 1 and d_cnt >= 1):
                contested.append(did)

        contested = sorted(set(contested))
        log.info("Спорных районов: %d (%s)", len(contested), contested)

        # Если нет спорных — можно сразу вернуть
        if not contested:
            return contested

        # Подготовим данные для уведомлений
        # 1) карта районов -> имя
        dq = await session.execute(
            select(District.id, District.name).where(District.id.in_(contested))
        )
        district_names = {int(i): n for i, n in dq.all()}

        # 2) список всех on-point действий в спорных районах
        notify_actions: list[tuple[int, int, int]] = []
        for did in contested:
            notify_actions.extend((aid, uid, did) for aid, uid in by_district[did]["attack"])
            notify_actions.extend((aid, uid, did) for aid, uid in by_district[did]["defense"])

        # 3) пользователи -> tg_id
        owner_ids = sorted({uid for _, uid, _ in notify_actions})
        uq = await session.execute(select(User.id, User.tg_id).where(User.id.in_(owner_ids)))
        users_map = {int(i): int(tg) for i, tg in uq.all() if tg is not None}

        # 4) пытаемся отправить интерактивный экран; если нет — шлём простое уведомление
        try:
            from app import bot  # type: ignore
        except Exception:
            bot = None
            log.warning("Бот недоступен: уведомления AskWhoWon пропущены.")

        AskWhoWon = None
        if bot:
            try:
                # поправьте путь импорта при необходимости
                from screens.notify_screen import AskWhoWonScreen  # type: ignore
                AskWhoWon = AskWhoWonScreen
            except Exception:
                AskWhoWon = None

        title = "⚔️ Спорный бой"
        # Рассылка
        if bot:
            for action_id, owner_id, did in notify_actions:
                tg_id = users_map.get(owner_id)
                if not tg_id:
                    continue
                body = f"В районе «{district_names.get(did, str(did))}» было несколько движений на точку. Вы победили?"
                try:
                    if AskWhoWon:
                        await AskWhoWon().run(
                            message=None,
                            actor=None,
                            state=None,
                            title=title,
                            body=body,
                            action_id=action_id,
                            bot=bot,
                            chat_id=tg_id
                        )
                    else:
                        await notify_user(bot, tg_id, title=title, body=body)
                except Exception:
                    log.exception("Не удалось отправить AskWhoWon (action_id=%s, owner_id=%s)", action_id, owner_id)

        return contested


# ===========================
#     DEFENSE POOLS
# ===========================
async def resolve_defense_pools(session: AsyncSession, rates: CombatRates, contested: List[int]) -> Dict[int, int]:
    """
    Формирует стартовый пул обороны из control_points (кроме спорных),
    затем добавляет очки из pending defense-действий (конверсия + on_point),
    и уведомляет владельца района о каждой защите.
    """
    with StepTimer("Формирование обороны"):
        # NEW: бот, если доступен
        try:
            from app import bot  # type: ignore
        except Exception:
            bot = None
            log.warning("Бот недоступен: уведомления о защите отправляться не будут.")

        contested_set = set(contested)

        # 0) Стартовая оборона из control_points
        q = await session.execute(select(District))
        districts: List[District] = list(q.scalars().all())
        log.info("Районов в базе: %d", len(districts))

        # NEW: быстрый доступ к району по id
        district_by_id = {d.id: d for d in districts}

        # NEW: кэш пользователей
        user_cache: Dict[int, User] = {}

        async def get_user(uid: int) -> Optional[User]:
            if uid not in user_cache:
                uq = await session.execute(select(User).where(User.id == uid))
                user_cache[uid] = uq.scalars().first()
            return user_cache.get(uid)

        defense_pool: Dict[int, int] = defaultdict(int)
        seeded_from_cp: Dict[int, int] = {}

        for d in districts:
            if d.id in contested_set:
                continue
            if d.control_points > 0:
                defense_pool[d.id] += int(d.control_points)
                seeded_from_cp[d.id] = int(d.control_points)
                d.control_points = 0

        if seeded_from_cp:
            await session.commit()
            log.info("Старт обороны из CP: %s", seeded_from_cp)

        # 1) Прибавляем оборону из pending defense-экшенов
        stmt = (
            select(Action)
            .where(
                Action.status == ActionStatus.PENDING,
                Action.kind == DEFENSE_KIND,
                Action.district_id.is_not(None),
            )
            .order_by(Action.id.asc())
        )
        res = await session.execute(stmt)
        actions: List[Action] = list(res.scalars().all())
        log.info("Активных защит: %d", len(actions))

        used_ids: List[int] = []
        for a in actions:
            did = a.district_id
            if did is None or did in contested_set:
                continue

            pts = resources_to_points(DEFENSE_KIND, a, rates)
            defense_pool[did] += pts
            used_ids.append(a.id)
            log.debug("DEF@%s: +%d очков (action #%s)", did, pts, a.id)

            # NEW: уведомление владельцу района о КАЖДОЙ защите
            if bot:
                d = district_by_id.get(did)
                if d:
                    owner = await get_user(d.owner_id)
                    defender = await get_user(a.owner_id)
                    if owner and owner.tg_id and defender:
                        defender_name = defender.in_game_name or defender.username or f"User#{defender.id}"
                        try:
                            await notify_user(
                                bot,
                                owner.tg_id,
                                title="🛡️ Район усилен защитой",
                                body=(
                                    f"Ваш район <b>{d.name}</b> защищён игроком <b>{defender_name}</b> "
                                    f"на <b>{pts}</b> очков контроля."
                                ),
                            )
                        except Exception:
                            log.exception("Не удалось отправить уведомление о защите (did=%s, action_id=%s)", did, a.id)

        if used_ids:
            await session.execute(
                update(Action)
                .where(Action.id.in_(used_ids))
                .values(status=ActionStatus.DONE, updated_at=now_utc())
            )
            await session.commit()

        log.info("Итоги обороны по районам (очки): %s", dict(defense_pool))
        return dict(defense_pool)


# ===========================
#     ATTACK RESOLUTION
# ===========================
async def resolve_attacks(session: AsyncSession, rates: CombatRates, defense_pool: Dict[int, int], contested: List[int]):
    """Пошагово резолвит атаки по районам, с учётом defense_pool и спорных районов."""
    with StepTimer("Разрешение атак"):
        # Бот для уведомлений (если есть)
        try:
            from app import bot  # type: ignore
        except Exception:
            bot = None
            log.warning("Бот недоступен: уведомления об атаках отправляться не будут.")

        base_stmt = (
            select(Action)
            .where(
                Action.status == ActionStatus.PENDING,
                Action.kind == ATTACK_KIND,
                Action.district_id.is_not(None),
            )
        )
        base_stmt = base_stmt.order_by(Action.created_at.asc() if ORDER_ATTACKS_ASC else Action.created_at.desc())
        res = await session.execute(base_stmt)
        attacks: List[Action] = list(res.scalars().all())
        log.info("Активных атак: %d", len(attacks))

        if not attacks:
            return

        by_district: Dict[int, List[Action]] = defaultdict(list)
        for a in attacks:
            if a.district_id:
                by_district[a.district_id].append(a)

        district_cache: Dict[int, District] = {}
        user_cache: Dict[int, User] = {}

        async def get_district(did: int) -> Optional[District]:
            if did not in district_cache:
                d = await District.get_by_id(session, did)
                if d:
                    district_cache[did] = d
                else:
                    return None
            return district_cache[did]

        async def get_user(uid: int) -> Optional[User]:
            if uid not in user_cache:
                q = await session.execute(select(User).where(User.id == uid))
                user = q.scalars().first()
                if user:
                    user_cache[uid] = user
                else:
                    return None
            return user_cache[uid]

        processed_ids: List[int] = []
        ownership_changes = 0

        for district_id, attack_list in by_district.items():
            if district_id in contested:
                log.info("Район %s спорный — атаки пропущены (%d шт.)", district_id, len(attack_list))
                continue

            d = await get_district(district_id)
            if not d:
                log.warning("Не найден район %s — пропущено %d атак", district_id, len(attack_list))
                for a in attack_list:
                    processed_ids.append(a.id)
                continue

            current_def = int(defense_pool.get(district_id, 0))
            log.info("Район '%s' стартовая оборона: %d", d.name, current_def)

            for a in attack_list:
                power_pts = resources_to_points(ATTACK_KIND, a, rates)
                attacker = await get_user(a.owner_id)
                attacker_name = (attacker.in_game_name or attacker.username or f"User#{attacker.id}") if attacker else "Неизвестный"
                attacker_faction = (attacker.faction or "без фракции") if attacker else "неизвестно"

                defender_user_before = await get_user(d.owner_id) if d.owner_id else None
                defender_name_before = (
                        defender_user_before.in_game_name
                        or defender_user_before.username
                        or (f"User#{defender_user_before.id}" if defender_user_before else "—")
                ) if defender_user_before else "—"
                def_before = current_def
                log.debug("ATK@%s by %s: %d pts vs def %d", d.id, attacker_name, power_pts, current_def)

                if power_pts <= current_def:
                    current_def -= power_pts
                    await add_news(
                        session,
                        title=f"Отражена атака на район '{d.name}'",
                        body=(
                            f"Атака игрока {attacker_name} ({power_pts} очков) была отражена. "
                            f"Фракция атакующего: {attacker_faction}. "
                            f"Текущая оборона района: {current_def}."
                        ),
                        action_id=a.id,
                        district_id=d.id,
                    )

                    if bot and attacker and defender_user_before:
                        await notify_user(
                            bot,
                            attacker.tg_id,
                            title="❌ Атака отражена",
                            body=(
                                f"Район <b>{d.name}</b> не взят. "
                                f"Ваши очки: <b>{power_pts}</b>. "
                                f"Оставшаяся оборона района: <b>{current_def}</b>."
                            ),
                        )
                        await notify_user(
                            bot,
                            defender_user_before.tg_id,
                            title="🛡️ Атака отражена",
                            body=(f"Ваш район <b>{d.name}</b> отбил атаку ({power_pts} очков). "
                                  f"Текущая оборона: <b>{current_def}</b>.")
                        )

                else:
                    overflow = power_pts - current_def
                    defender_user = defender_user_before

                    d = await District.reassign_owner(session, district_id=district_id, new_owner_id=a.owner_id)
                    district_cache[district_id] = d
                    current_def = overflow
                    ownership_changes += 1

                    await add_news(
                        session,
                        title=f"Район '{d.name}' захвачен!",
                        body=(
                            f"Атака игрока {attacker_name} ({power_pts} очков) прорвала оборону района. "
                            f"Фракция захватившего: {attacker_faction}. "
                            f"Новый владелец — {attacker_name}. Остаток {overflow} очков укрепил оборону района."
                        ),
                        action_id=a.id,
                        district_id=d.id,
                    )

                    if bot and attacker and defender_user:
                        await notify_user(
                            bot,
                            attacker.tg_id,
                            title="✅ Район захвачен",
                            body=(
                                f"Вы захватили район <b>{d.name}</b>! "
                                f"Прорыв силой <b>{power_pts}</b>. "
                                f"Остаток <b>{overflow}</b> стал обороной района."
                            ),
                        )
                        await notify_user(
                            bot,
                            defender_user.tg_id,
                            title="⚠️ Потеря района",
                            body="Ваш район <b>{}</b> был атакован {} и утерян.".format(d.name, attacker_name),
                        )
                    try:
                        raw_body = (
                            f'Бой за район "{d.name}". '
                            f'Нападающий: "{attacker_name}". '
                            f'Оборонявшийся: "{defender_name_before}". '
                            f'Победил: "{attacker_name}". '
                            f'Силы: атака {power_pts} против обороны {def_before}. '
                            f'Остаток {overflow} пошёл в оборону района.'
                        )
                        await asyncio.to_thread(add_raw_row, raw_body=raw_body, type_value="battle")
                    except Exception:
                        log.exception("Не удалось записать RAW протокол боя (district_id=%s, action_id=%s)",
                                      district_id, a.id)
                processed_ids.append(a.id)

            defense_pool[district_id] = current_def
            log.info("Район '%s' остаточная оборона после атак: %d", d.name, current_def)

        if processed_ids:
            await session.execute(
                update(Action).where(Action.id.in_(processed_ids)).values(status=ActionStatus.DONE, updated_at=now_utc())
            )
            await session.commit()
            log.info("Атак обработано и закрыто: %d", len(processed_ids))

        if ownership_changes:
            log.info("Смен владельцев районов: %d", ownership_changes)


# ===========================
#  LEFTOVER DEF → CONTROL
# ===========================
async def convert_leftover_defense_to_control_points(
    session: AsyncSession,
    defense_pool: Dict[int, int],
    contested: List[int],
) -> None:
    """Остаток обороны → в control_points (спорные районы пропускаются)."""
    with StepTimer("Конвертация остатка обороны в CP"):
        if not defense_pool:
            log.info("Defense pool пуст — нечего конвертировать.")
            return

        contested_set = set(contested)
        updated: Dict[int, int] = {}

        for district_id, remaining_def in defense_pool.items():
            if remaining_def <= 0 or district_id in contested_set:
                continue
            d = await District.get_by_id(session, district_id)
            if not d:
                continue
            d.control_points += int(remaining_def)
            updated[district_id] = int(remaining_def)

        if updated:
            await session.commit()
            log.info("В CP начислено: %s", updated)
        else:
            log.info("Начислений в CP нет.")


# ===========================
#     CLOSE ALL SCOUTING
# ===========================
async def close_all_scouting(session: AsyncSession):
    """
    Закрывает все pending-разведки, сбрасывает связи scout (User <-> District)
    и уведомляет пользователей, наблюдавших за районами.
    """
    with StepTimer("Закрытие разведок и сброс наблюдения"):
        # Бот для уведомлений (если есть)
        try:
            from app import bot  # type: ignore
        except Exception:
            bot = None
            log.warning("Бот недоступен: уведомления о завершении разведки отправляться не будут.")

        rows = await session.execute(
            select(
                user_scouts_districts.c.user_id,
                user_scouts_districts.c.district_id,
                District.name,
            ).join(District, District.id == user_scouts_districts.c.district_id)
        )
        watched: Dict[int, List[tuple[int, str]]] = defaultdict(list)
        for uid, did, dname in rows.all():
            watched[int(uid)].append((int(did), dname))
        log.info("Связей наблюдения перед сбросом: %d", sum(len(v) for v in watched.values()))

        # Закрываем все PENDING scout-экшены
        stmt = select(Action.id).where(
            Action.status == ActionStatus.PENDING,
            Action.kind.in_(SCOUT_KINDS),
        )
        res = await session.execute(stmt)
        ids = [row for row in res.scalars().all()]
        if ids:
            await session.execute(
                update(Action).where(Action.id.in_(ids)).values(status=ActionStatus.DONE, updated_at=now_utc())
            )
            log.info("Закрыто разведывательных экшенов: %d", len(ids))

        # Сбрасываем связи наблюдения
        await session.execute(delete(user_scouts_districts))
        await session.commit()
        log.info("Связи наблюдения очищены.")

        # Уведомления
        if bot and watched:
            user_ids = list(watched.keys())
            q = await session.execute(select(User).where(User.id.in_(user_ids)))
            users_map = {u.id: u for u in q.scalars().all()}
            for uid, items in watched.items():
                user = users_map.get(uid)
                if not user or not items:
                    continue
                lines = [f"• {name} (#{did})" for did, name in items]
                body = (
                    "Разведка районов завершена, наблюдение сброшено.\n\n"
                    "Список районов, за которыми вы наблюдали:\n" + "\n".join(lines) +
                    "\n\nЧтобы продолжить наблюдение, запустите новую разведку."
                )
                await notify_user(bot, user.tg_id, title="🔍 Разведка завершена", body=body)


# ===========================
#  IDEOLOGY MULTIPLIER
# ===========================
def ideology_multiplier(owner_ideol: int, pol_ideol: Optional[int]) -> float:
    """
    diff=|owner - politician| ∈ [0..10]
      diff=0  -> 1.20
      diff=10 -> 0.40
      линейно, шаг 0.08 за единицу diff.
    Если политика нет — возвращаем 1.0 (и multiplier не меняем).
    """
    if pol_ideol is None:
        return 1.0
    diff = abs(int(owner_ideol) - int(pol_ideol))  # 0..10
    diff = max(0, min(10, diff))
    return 1.20 - 0.08 * diff


async def recalc_resource_multipliers(session: AsyncSession):
    """Обновляет district.resource_multiplier (кладём в шаг 0.1)."""
    with StepTimer("Пересчёт множителей ресурсов по идеологии"):
        q = await session.execute(select(District))
        districts: List[District] = list(q.scalars().all())
        if not districts:
            log.info("Районов нет — пересчитывать нечего.")
            return

        owner_cache: Dict[int, User] = {}
        pol_by_district: Dict[int, Optional[Politician]] = {}

        for d in districts:
            if d.owner_id not in owner_cache:
                uq = await session.execute(select(User).where(User.id == d.owner_id))
                owner_cache[d.owner_id] = uq.scalars().first()

            pq = await session.execute(select(Politician).where(Politician.district_id == d.id))
            pol = pq.scalars().first()
            pol_by_district[d.id] = pol

        updated = 0
        for d in districts:
            owner = owner_cache.get(d.owner_id)
            pol = pol_by_district.get(d.id)
            if owner and pol:
                mul = ideology_multiplier(owner.ideology, pol.ideology)
                mul = _quantize_tenth(mul, 0.40, 1.20)
                if abs(d.resource_multiplier - mul) > 1e-6:
                    log.debug(
                        "District %s: mul %.2f → %.2f (owner=%s, pol=%s)",
                        d.id, d.resource_multiplier, mul, owner.ideology, pol.ideology
                    )
                    d.resource_multiplier = mul
                    updated += 1

        if updated:
            await session.commit()
            log.info("Обновлены множители ресурсов у %d районов.", updated)
        else:
            log.info("Множители ресурсов без изменений.")


# ===========================
#  GRANT USERS' BASE RESOURCES
# ===========================
async def grant_users_base_resources(session: AsyncSession):
    """
    Начисляет каждому пользователю его базовые ресурсы (user.base_*).
    Базовые ресурсы НЕ умножаются и просто добавляются к накопленным.
    Для каждого с ненулевыми базовыми — отправляется нотификация (если доступен bot).
    """
    with StepTimer("Начисление базовых ресурсов игрокам"):
        # Бот для уведомлений (если есть)
        try:
            from app import bot  # type: ignore
        except Exception:
            bot = None
            log.warning("Бот недоступен: уведомления о базовых ресурсах отправляться не будут.")

        res = await session.execute(select(User))
        users: List[User] = list(res.scalars().all())
        if not users:
            log.info("Пользователей нет — базовые ресурсы начислять некому.")
            return

        total_money = total_infl = total_info = total_force = 0
        # Запомним, кому и что начислили — нотифицируем после commit
        to_notify: List[tuple[int, int, dict]] = []  # (user_id, user_tg_id, delta_dict)

        for u in users:
            bm  = max(0, int(u.base_money or 0))
            bi  = max(0, int(u.base_influence or 0))
            binf = max(0, int(u.base_information or 0))
            bf  = max(0, int(u.base_force or 0))

            if bm == 0 and bi == 0 and binf == 0 and bf == 0:
                continue

            u.money       += bm
            u.influence   += bi
            u.information += binf
            u.force       += bf

            total_money += bm
            total_infl  += bi
            total_info  += binf
            total_force += bf

            to_notify.append((u.id, u.tg_id, {
                "money": bm, "influence": bi, "information": binf, "force": bf
            }))

        await session.commit()
        log.info(
            "Базовые ресурсы начислены суммарно: 💰%s 🪙%s 🧠%s 💪%s (получателей: %d)",
            total_money, total_infl, total_info, total_force, len(to_notify)
        )

        # Нотификации
        if bot and to_notify:
            # подгрузим имена только для тех, кому начисляли (опционально)
            user_ids = [uid for uid, _, _ in to_notify]
            q = await session.execute(select(User).where(User.id.in_(user_ids)))
            users_map = {u.id: u for u in q.scalars().all()}

            for uid, tg_id, delta in to_notify:
                # если вдруг пользователя уже нет — пропускаем
                user = users_map.get(uid)
                if not user:
                    continue

                body = (
                    "Вам начислены базовые ресурсы:\n"
                    f"• 💰 {delta['money']}\n"
                    f"• 🪙 {delta['influence']}\n"
                    f"• 🧠 {delta['information']}\n"
                    f"• 💪 {delta['force']}\n"
                )
                try:
                    await notify_user(
                        bot,
                        tg_id,
                        title="📦 Базовые ресурсы начислены",
                        body=body,
                    )
                except Exception:
                    # не валим цикл из-за одного неотправленного сообщения
                    log.exception("Не удалось отправить нотификацию о базовых ресурсах пользователю #%s", uid)

async def process_politician_influence(session: AsyncSession) -> None:
    """Обрабатывает pending-заявки вида 'influence' и меняет идеологию политиков."""
    with StepTimer("Обработка влияния на политиков"):
        # Бот для уведомлений (если есть)
        try:
            from app import bot  # type: ignore
        except Exception:
            bot = None
            log.warning("Бот недоступен: уведомления о влиянии отправляться не будут.")

        # 1) Собираем заявки influence, старые → новые
        stmt = (
            select(Action)
            .where(
                Action.status == ActionStatus.PENDING,
                Action.kind == "influence",     # ключевой признак
                # при желании можно усилить фильтр:
                # Action.type == ActionType.INFLUENCE
            )
            .order_by(Action.created_at.asc())
        )
        res = await session.execute(stmt)
        acts: List[Action] = list(res.scalars().all())
        log.info("Активных influence-заявок: %d", len(acts))
        if not acts:
            return

        # 2) Кэш политиков по районам
        district_ids = sorted({a.district_id for a in acts if a.district_id is not None})
        pol_by_district: Dict[int, Politician] = {}
        pol_by_id: Dict[int, Politician] = {}
        ideology_map: Dict[int, int] = {}

        if district_ids:
            pq = await session.execute(
                select(Politician).where(Politician.district_id.in_(district_ids))
            )
            pols = list(pq.scalars().all())
            for p in pols:
                if p.district_id is not None:
                    pol_by_district[p.district_id] = p
                pol_by_id[p.id] = p
                ideology_map[p.id] = int(p.ideology or 0)

        processed_ids: List[int] = []
        changed_pids: set[int] = set()
        notify_pairs: set[tuple[int, int]] = set()  # (user_id, politician_id)

        # 3) Применяем влияние к temp-идеологии
        for a in acts:
            did = a.district_id
            p = pol_by_district.get(did) if did is not None else None
            processed_ids.append(a.id)  # заявку закрываем в любом случае

            if not p:
                log.debug("Influence action #%s: нет политика для района %s", a.id, did)
                continue

            amt = max(0, int(a.influence or 0))
            cur_val = ideology_map.get(p.id, int(p.ideology or 0))

            if a.is_positive is True:
                cur_val += amt
            elif a.is_positive is False:
                cur_val -= amt
            else:
                log.debug("Influence action #%s: is_positive=None, изменение пропущено", a.id)

            ideology_map[p.id] = cur_val
            changed_pids.add(p.id)
            notify_pairs.add((a.owner_id, p.id))

        # 4) Закрываем заявки
        if processed_ids:
            await session.execute(
                update(Action)
                .where(Action.id.in_(processed_ids))
                .values(status=ActionStatus.DONE, updated_at=now_utc())
            )
            await session.commit()
            log.info("Закрыто influence-заявок: %d", len(processed_ids))

        # 5) Квантуем идеологию в [-5..5] и сохраняем в БД
        if changed_pids:
            for pid in changed_pids:
                p = pol_by_id.get(pid)
                if not p:
                    continue
                new_val = ideology_map.get(pid, p.ideology)
                new_val = max(-5, min(5, int(new_val)))
                if p.ideology != new_val:
                    log.debug("Политик #%s (%s): идеология %s → %s", p.id, p.name, p.ideology, new_val)
                    p.ideology = new_val

            await session.commit()
            log.info("Обновлена идеология у %d политиков.", len(changed_pids))

        # 6) Уведомления авторам заявок
        if bot and notify_pairs:
            user_ids = sorted({uid for uid, _ in notify_pairs})
            uq = await session.execute(select(User).where(User.id.in_(user_ids)))
            users_map = {u.id: u for u in uq.scalars().all()}

            for uid, pid in sorted(notify_pairs):
                user = users_map.get(uid)
                pol = pol_by_id.get(pid)
                if not user or not pol:
                    continue
                try:
                    await notify_user(
                        bot,
                        user.tg_id,
                        title="🏛️ Влияние учтено",
                        body=f"Ваши действия повлияли на политика «{pol.name}».",
                    )
                except Exception:
                    log.exception("Не удалось отправить уведомление о влиянии (user_id=%s, pol_id=%s)", uid, pid)

# ===========================
#    GRANT RESOURCES
# ===========================
async def grant_district_resources(session: AsyncSession, contested: List[int]):
    """Начисляет владельцам ресурсы с районов, исключая спорные (и шлёт уведомления)."""
    with StepTimer("Начисление ресурсов районам"):
        # Бот для уведомлений (если есть)
        try:
            from app import bot  # type: ignore
        except Exception:
            bot = None
            log.warning("Бот недоступен: уведомления о начислении ресурсов отправляться не будут.")

        res = await session.execute(select(District))
        districts: List[District] = list(res.scalars().all())
        if not districts:
            log.info("Районов нет — начислять нечего.")
            return

        contested_set = set(contested)
        changes: Dict[int, dict] = defaultdict(lambda: {"money": 0, "influence": 0, "information": 0, "force": 0})
        per_owner_breakdown: Dict[int, List[tuple[str, dict]]] = defaultdict(list)

        for d in districts:
            if d.id in contested_set:
                log.debug("Район '%s' спорный — пропуск начисления.", d.name)
                continue
            eff = d.effective_resources()
            changes[d.owner_id]["money"] += eff["money"]
            changes[d.owner_id]["influence"] += eff["influence"]
            changes[d.owner_id]["information"] += eff["information"]
            changes[d.owner_id]["force"] += eff["force"]

            per_owner_breakdown[d.owner_id].append(
                (d.name, {"money": eff["money"], "influence": eff["influence"], "information": eff["information"], "force": eff["force"]})
            )

        total_money = total_infl = total_info = total_force = 0

        for uid, delta in changes.items():
            uq = await session.execute(select(User).where(User.id == uid))
            user = uq.scalars().first()
            if not user:
                continue

            user.money += delta["money"]
            user.influence += delta["influence"]
            user.information += delta["information"]
            user.force += delta["force"]

            total_money += delta["money"]
            total_infl += delta["influence"]
            total_info += delta["information"]
            total_force += delta["force"]

        await session.commit()
        log.info("Начислено суммарно: 💰%s 🪙%s 🧠%s 💪%s", total_money, total_infl, total_info, total_force)

        # Уведомления
        if bot:
            for uid, items in per_owner_breakdown.items():
                sums = changes.get(uid, {})
                if not sums or (sums["money"] + sums["influence"] + sums["information"] + sums["force"]) <= 0:
                    continue

                uq = await session.execute(select(User).where(User.id == uid))
                user = uq.scalars().first()
                if not user:
                    continue

                lines = []
                for name, r in items:
                    lines.append(
                        f"• <b>{name}</b>: 💰 {r['money']}, 🪙 {r['influence']}, 🧠 {r['information']}, 💪 {r['force']}"
                    )
                total_line = (
                    f"Итого: 💰 <b>{sums['money']}</b>, 🪙 <b>{sums['influence']}</b>, "
                    f"🧠 <b>{sums['information']}</b>, 💪 <b>{sums['force']}</b>"
                )
                body = "Вы получили ресурсы с контролируемых районов:\n" + "\n".join(lines) + "\n\n" + total_line

                await notify_user(
                    bot,
                    user.tg_id,
                    title="💼 Ресурсы начислены",
                    body=body,
                )


# ===========================
#    REFRESH ACTION SLOTS
# ===========================
async def refresh_player_actions(session: AsyncSession):
    """Восстанавливает слоты действий игрокам до максимума."""
    with StepTimer("Обновление слотов действий"):
        res = await session.execute(select(User))
        users: List[User] = list(res.scalars().all())
        if not users:
            log.info("Пользователей нет — слоты обновлять некому.")
            return

        refreshed = 0
        for u in users:
            max_actions = u.max_available_actions or 0
            if u.available_actions < max_actions:
                u.available_actions = max_actions
                u.actions_refresh_at = now_utc()
                refreshed += 1

        await session.commit()
        log.info("Слоты действий восстановлены у %d игроков.", refreshed)


# ===========================
#           MAIN
# ===========================
async def run_game_cycle():
    global CYCLE_TS

    # 0) загрузить курсы конверсии
    with StepTimer("Инициализация курсов"):
        rates = CombatRates.load(COMBAT_RATES_PATH)

    # фиксируем timestamp цикла и создаём файл
    CYCLE_TS = now_utc().strftime("%Y%m%dT%H%M%SZ")
    _ensure_cycle_workbook()
    log.info("Таймстемп цикла (UTC): %s", CYCLE_TS)

    engine = create_async_engine(DATABASE_URL, echo=False, future=True)
    async_session_factory = sessionmaker(engine, expire_on_commit=False, class_=AsyncSession)

    async with async_session_factory() as session:
        log.info("=== Старт игрового цикла ===")

        try:
            with StepTimer("Шаг A: Агрегация SUPPORT"):
                await aggregate_supports(session)

            with StepTimer("Шаг B: Определение спорных районов"):
                contested = await detect_contested_districts(session)

            with StepTimer("Шаг 1: Резерв обороны"):
                defense_pool = await resolve_defense_pools(session, rates, contested)

            with StepTimer("Шаг 2: Резолв атак"):
                await resolve_attacks(session, rates, defense_pool, contested)

            with StepTimer("Шаг 2.5: Остаток обороны → CP"):
                await convert_leftover_defense_to_control_points(session, defense_pool, contested)

            with StepTimer("Шаг 2.6: Влияние на политиков"):
                await process_politician_influence(session)

            with StepTimer("Шаг 3: Закрыть все разведки"):
                await close_all_scouting(session)

            with StepTimer("Шаг 4: Пересчёт ресурсных множителей"):
                await recalc_resource_multipliers(session)

            with StepTimer("Шаг 4.5: Базовые ресурсы игрокам"):
                await grant_users_base_resources(session)

            with StepTimer("Шаг 5: Выдача ресурсов"):
                await grant_district_resources(session, contested)

            with StepTimer("Шаг 6: Обновление слотов действий"):
                await refresh_player_actions(session)

            log.info("=== Игровой цикл завершён ===")
            try:
                Path("last_cycle_finished.txt").write_text(now_utc().isoformat(), encoding="utf-8")
                log.info("Записан маркер завершения цикла: last_cycle_finished.txt")
            except Exception:
                log.exception("Не удалось записать last_cycle_finished.txt")

        except Exception:
            log.exception("Игровой цикл завершился с ошибкой")
            raise


if __name__ == "__main__":
    asyncio.run(run_game_cycle())
